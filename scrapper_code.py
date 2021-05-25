import requests
from bs4 import BeautifulSoup as Soup
import openpyxl as xl
from openpyxl import Workbook
from openpyxl import load_workbook
import time

# selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import selenium.webdriver.support.ui as ui


links = 'dismay.html'
DRIVER_PATH = 'chromedriver'
dismay_categories = {}
brasilybelleza_categories = {}

sublime_bw_categories = {'CABELLO - Extensiones': 'https://sublimebw.com/4489-extensiones',
                         'Estetica': 'https://sublimebw.com/162-estetica',
                         'Utillaje': 'https://sublimebw.com/148-utillaje'}

headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64; rv:60.0) Gecko/20100101 Firefox/60.0",
           "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
           "Accept-Language": "en-US,en;q=0.9"
           }
# save data for brasilybelleza
wb_brasilybelleza = Workbook()
ws_brasilybelleza = wb_brasilybelleza.active
product_brasilybelleza = 'product_brasilybelleza.xlsx'
sheet_title = 'products'
ws_brasilybelleza.append(["Title", "Brand", "Category", "Image", "Short Description", "Detail Description"])
wb_brasilybelleza.save(product_brasilybelleza)

# save data for dismay
wb_dismay = Workbook()
ws_dismay = wb_dismay.active
product_dismay = 'product_dismay.xlsx'
sheet_title = 'products'
ws_dismay.append(["Title", "Brand", "Category", "Image", "Short Description", "Detail Description"])
wb_dismay.save(product_dismay)

# save worksheet for sublime_bw
wb_sublime_bw = Workbook()
ws_sublime_bw = wb_sublime_bw.active
product_sublime_bw = 'product_sublime_bw.xlsx'
sheet_title = 'products'
ws_sublime_bw.append(["Title", "Brand", "Category", "Image", "Short Description", "Detail Description"])
wb_sublime_bw.save(product_sublime_bw)


# this method will scrape dismay data
def get_dismay():
    # get all categories
    link_url = "https://shop.dismay.es/"
    response_ = requests.get(url=link_url, headers=headers)
    data_soup = Soup(response_.text, "html.parser")
    print(data_soup)
    categories = data_soup.find('ul', {'class:', 'list-block list-group bullet tree dhtml'})
    for category in categories.findAll('li'):
        cat_name = category.find('a').text
        cat_url = category.find('a')['href']
        # save categories to dictionary
        dismay_categories[cat_name] = cat_url

    time.sleep(2)
    # getting each category
    for key in dismay_categories:
        url_length = 0
        time.sleep(5)
        print(key + ' : ' + dismay_categories[key] + '\n\n')
        cat_url = dismay_categories[key]
        cat_response = requests.get(url=cat_url, headers=headers)
        # checking for subcategories
        if 'subcategories' in cat_response.text:
            print("its have sub")
        else:
            # get all pages in a category
            cat_soup = Soup(cat_response.text, "html.parser")
            try:
                cat_pages = cat_soup.find('ul', {'class:', 'pagination pull-left'})
                if cat_pages is None:
                    url_length = 1
                else:
                    page_length = len(cat_pages.findAll('a'))
                    page_length = link_url + cat_pages.findAll('a')[page_length-2]['href']
                    url_length = len(page_length)
                    url_length = page_length[url_length-1]
            except:
                url_length = 1

        # get product page url for each product in a page
        try:
            for page_ in range(int(url_length)):
                page_url = cat_url + '?p=' + str(page_ + 1)
                time.sleep(2)
                print("Page URL: " + page_url)
                products_response = requests.get(url=page_url, headers=headers)
                products_soup = Soup(products_response.text, "html.parser")
                products = products_soup.find('div', {'class:', 'product_list grid row'})
                for product in products.findAll('div', {'class:', 'product-container'}):
                    product_url = product.find('a', {'class:', 'product_img_link'})
                    time.sleep(1)
                    # product details page / detail page url
                    d_url = product_url['href']
                    products_details = requests.get(url=d_url, headers=headers)
                    details_soup = Soup(products_details.text, "html.parser")
                    all_category = details_soup.find('span', {'class:', 'navigation_page'})

                    # check if product in sub category
                    check_sub_category = len(all_category.findAll('span'))
                    if check_sub_category > 3:
                        category_name = all_category.findAll('span')[0].text
                        sub_cat_name = all_category.findAll('span')[3].text

                    else:
                        category_name = all_category.findAll('span')[0].text
                        sub_cat_name = ''

                    product_section = details_soup.find('div', {'class:', 'primary_block row'})

                    # product image
                    image_src = product_section.find('img')['src']

                    # product name
                    product_name = product_section.find('h1').text

                    # Brand and Reference
                    try:
                        for brand_reference in product_section.findAll('p'):
                            if 'Marca' in brand_reference.text or 'Brand' in brand_reference.text:
                                brand_name = brand_reference.text
                            if 'Referencia' in brand_reference.text or 'Reference' in brand_reference.text:
                                reference_name = brand_reference.text
                    except:
                        brand_name = ''
                        reference_name = ''

                    # short description
                    try:
                        short_description = product_section.find('div', {'id': 'short_description_content'})
                        short_description = short_description.text
                    except:
                        short_description = ''

                    # detail description
                    try:
                        long_description = details_soup.find('div', {'class:', 'more_info_block'})
                        long_description = long_description.find('div', {'class:', 'tab-content'})
                        long_description = long_description.text
                    except:
                        long_description = ''

                    ws_dismay.append(
                        [product_name, brand_name, category_name, image_src, short_description, long_description])
                    wb_dismay.save(product_dismay)
        except:
            pass


# method for scrapping sublime data
def get_sublime_bw():
    # get all categories
    for key in sublime_bw_categories:
        category_name = key
        print(key + ' : ' + sublime_bw_categories[key])
        link_url = sublime_bw_categories[key]
        response_ = requests.get(url=link_url, headers=headers)
        data_soup = Soup(response_.text, "html.parser")

        pages = data_soup.find('ul', {'class:', 'page-list clearfix text-md-right text-xs-center'})
        page_length = len(pages.findAll('li'))
        page_length = pages.findAll('li')[page_length-2].text

        # go to each page inside category
        for i in range(int(page_length)):
            page_link = link_url + "?page="+str(i+1)
            print("page: " + page_link)
            page_response_ = requests.get(url=page_link, headers=headers)
            page_soup = Soup(page_response_.text, "html.parser")
            products = page_soup.find('div', {'class:', 'products'})
            products = products.find('div', {'class:', 'product_list grid plist-default'})
            products = products.find('div', {'class:', 'row'})
            for product in products.findAll('div', {'class:', 'ajax_block_product'}):
                product_page_link = product.find('a')['href']
                try:
                    product_response_ = requests.get(
                        url=product_page_link,
                        headers=headers)
                    product_soup = Soup(product_response_.text, "html.parser")
                    product_soup = product_soup.find('section', {'id': 'wrapper'})
                    nav_bar = product_soup.find('nav', {'class:', 'breadcrumb hidden-sm-down'})
                    nav_bar = nav_bar.find('ol')

                    # brand name
                    brand = nav_bar.findAll('li')[2].text.strip()

                    main_section = product_soup.find('section', {'id': 'main'})
                    main_section = main_section.find('div', {'class:', 'row'})

                    product_image = main_section.find('img', {'id': 'zoom_product'})['src']
                    product_name = main_section.find('h1', {'class:', 'h1 product-detail-name'}).text
                    product_manufacturer = brand
                    product_reference = main_section.find('div', {'class:', 'product-reference'}).text.strip()

                    ws_sublime_bw.append([product_name, brand, category_name, product_image, '', ''])
                    wb_sublime_bw.save(product_sublime_bw)
                except:
                    pass


# method for scrapping brasily data
def get_brasily_belleza():

    # PART 1 here......
    link_url = 'https://www.brasilybelleza.com'
    response_ = requests.get(url=link_url, headers=headers)
    data_soup = Soup(response_.text, "html.parser")
    categories = data_soup.find('ul', {'class:', 'list-unstyled components'})
    for category in categories.findAll('li'):
        cat_key = category.find('a').text.strip()
        cat_link = link_url + category.find('a')['href']
        brasilybelleza_categories[cat_key] = cat_link

    for key in brasilybelleza_categories:
        print(key + ' : ' + brasilybelleza_categories[key])
        _id = brasilybelleza_categories[key].split('/')[6]
        print(_id)
        if 'Ofertas hasta -50%' in key or 'Tratamiento Queratina' in key or 'Alisado Brasileño' in key \
                or 'Botox Capilar' in key or 'Champú Antiresiduos' in key or 'Champú Sin Sal' in key \
                or 'Champú Sin Sulfatos' in key or 'Mascarilla Capilar' in key:
            print('already...')
            pass
        else:
            category_link = brasilybelleza_categories[key]
            options = Options()
            options.headless = True
            options.add_argument("--window-size=1920,1200")
            driver = webdriver.Chrome(options=options, executable_path=DRIVER_PATH)
            driver.get(category_link)
            time.sleep(1)
            html = driver.page_source
            time.sleep(2)
            while 'Ver más' in html:
                li_click = ActionChains(driver)
                li_click.click(driver.find_element_by_id(_id)).perform()
                time.sleep(10)
                html = driver.page_source
            data_soup = Soup(html, 'html.parser')
            products = data_soup.find('div', {'class:', 'row more_results_container'})
            try:
                for product in products.findAll('div', {'class:', 'col-md-3 col-sm-6 col-xs-6'}):
                    product_detail_link = link_url + product.find('a')['href']
                    print(product_detail_link)
                    try:
                        # get product details
                        product_url = product_detail_link
                        response_ = requests.get(url=product_url, headers=headers)
                        data_soup = Soup(response_.text, "html.parser")

                        # product details
                        product_name = data_soup.find('h1', {'id': 'product-name'}).text
                        product_image = data_soup.find('img', {'id': 'img-1'})['src']
                        short_description = data_soup.find('div', {'class:', 'col-md-6 col-sm-6 col-xs-12 product'}).text
                        short_description = short_description.replace(product_name, '').split('Referencia:', 1)[0].strip()

                        meta_datas = data_soup.find('ul', {'class:', 'product-list'})
                        reference = meta_datas.findAll('li')[0].text.strip()
                        brand = meta_datas.findAll('li')[1].text.strip()
                        availability = meta_datas.findAll('li')[2].text.strip()

                        price = data_soup.find('div', {'class:', 'price-box-price'}).text.strip()
                        price_weight = data_soup.find('div', {'class:', 'price-weight'}).text.strip()
                        full_description = data_soup.find('div', {'id': 'product-description'}).text.strip()

                        ws_brasilybelleza.append(
                            [product_name, brand, key, product_image, short_description, full_description])
                        wb_brasilybelleza.save(product_brasilybelleza)
                        time.sleep(2)
                    except:
                        pass
            except:
                pass


if __name__ == '__main__':

    get_dismay
    get_sublime_bw()
    get_brasily_belleza()
